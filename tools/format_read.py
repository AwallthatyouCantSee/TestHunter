# -*- coding: utf-8 -*-
"""FormatRead - 多格式文件读取工具

将 xlsx/docx/pdf/json/csv/yaml 等格式转为可读文本，解决内置 Read 工具
只能读 UTF-8 纯文本的限制。
"""
import csv
import io
import json
from pathlib import Path
from typing import Any

import aiofiles

from agentscope.tool import ToolBase
from agentscope.tool._response import ToolChunk
from agentscope.message import TextBlock, ToolResultState
from agentscope.permission import (
    PermissionContext,
    PermissionDecision,
    PermissionBehavior,
)
from agentscope.state import AgentState


class FormatRead(ToolBase):
    """多格式文件读取工具

    与 Read 工具的区别：Read 只能以 UTF-8 模式读取纯文本，
    此工具可解析二进制/结构化格式并转为 Markdown 文本。
    """

    name: str = "FormatRead"
    """工具名称"""

    # pylint: disable=line-too-long
    description: str = """Read xlsx/docx/pdf/json/csv/yaml files and convert them to readable text.

Unlike the Read tool (which only reads plain text in UTF-8 mode), this tool can parse structured/binary formats:
- **.xlsx**: Converts each sheet to a Markdown table
- **.docx**: Extracts paragraph text
- **.pdf**: Extracts text from all pages
- **.json**: Pretty-prints with indentation
- **.csv**: Converts to Markdown table
- **.yaml/.yml**: Outputs in structured YAML format
- Other text formats (.md/.py/.js/.sql/.html etc.): Reads as plain text"""
    """工具描述"""

    input_schema: dict[str, Any] = {
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "description": (
                    "The absolute path to the file to read. "
                    "Must be an absolute path, not relative."
                ),
            },
        },
        "required": ["file_path"],
    }
    """输入参数 schema"""

    is_mcp: bool = False
    is_read_only: bool = True
    is_concurrency_safe: bool = True
    is_external_tool: bool = False

    def __init__(self, max_table_rows: int = 500) -> None:
        """Initialize.

        Args:
            max_table_rows: Max rows to read from xlsx/csv (prevents
                blowing up context).
        """
        self._max_table_rows = max_table_rows

    async def check_permissions(
        self,
        tool_input: dict[str, Any],
        context: PermissionContext,
    ) -> PermissionDecision:
        """Read-only tool, always passthrough."""
        return PermissionDecision(
            behavior=PermissionBehavior.PASSTHROUGH,
            message="FormatRead is read-only.",
        )

    async def __call__(
        self,
        file_path: str,
        _agent_state: AgentState | None = None,
    ) -> ToolChunk:
        """Read and convert the file."""

        path = Path(file_path)

        # Validate absolute path
        if not path.is_absolute():
            return ToolChunk(
                content=[
                    TextBlock(
                        text=f"Error: file_path must be absolute, got: "
                        f"{file_path}",
                    ),
                ],
                state=ToolResultState.ERROR,
                is_last=True,
            )

        # Check exists
        if not path.exists():
            return ToolChunk(
                content=[
                    TextBlock(
                        text=f"Error: file does not exist: {file_path}",
                    ),
                ],
                state=ToolResultState.ERROR,
                is_last=True,
            )

        if path.is_dir():
            return ToolChunk(
                content=[
                    TextBlock(
                        text=f"Error: path is a directory: {file_path}",
                    ),
                ],
                state=ToolResultState.ERROR,
                is_last=True,
            )

        suffix = path.suffix.lower()

        try:
            text = await self._parse(path, suffix)
            return ToolChunk(
                content=[TextBlock(text=text)],
                state=ToolResultState.RUNNING,
                is_last=True,
            )
        except Exception as e:
            return ToolChunk(
                content=[
                    TextBlock(text=f"Error reading file: {e}"),
                ],
                state=ToolResultState.ERROR,
                is_last=True,
            )

    # ==================================================================
    # Format dispatchers
    # ==================================================================

    async def _parse(self, path: Path, suffix: str) -> str:
        """Dispatch to the appropriate parser."""
        if suffix == ".xlsx":
            return self._parse_xlsx(path)
        if suffix == ".docx":
            return self._parse_docx(path)
        if suffix == ".pdf":
            return self._parse_pdf(path)
        if suffix in (".json",):
            return await self._parse_json(path)
        if suffix in (".csv",):
            return await self._parse_csv(path)
        if suffix in (".yaml", ".yml"):
            return await self._parse_yaml(path)
        # Fallback: plain text
        return await self._parse_text(path)

    def _parse_xlsx(self, path: Path) -> str:
        """Convert xlsx sheets to Markdown tables."""
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)
        parts: list[str] = []
        for name in wb.sheetnames:
            ws = wb[name]
            parts.append(f"## Sheet: {name}\n")
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                parts.append("*(empty sheet)*\n")
                continue

            # Determine actual column count (ignore trailing None-only cols)
            col_count = max(
                (
                    max(
                        i + 1
                        for i, c in enumerate(row)
                        if c is not None
                    )
                    if row
                    else 0
                )
                for row in rows
            )
            col_count = max(col_count, 1)

            shown = 0
            for row in rows:
                if shown >= self._max_table_rows:
                    break
                cells = [
                    str(c) if c is not None else ""
                    for c in (list(row[:col_count]) if row else [])
                ]
                parts.append("| " + " | ".join(cells) + " |")
                shown += 1

            if len(rows) > self._max_table_rows:
                parts.append(
                    f"\n*(共 {len(rows)} 行，仅展示前 "
                    f"{self._max_table_rows} 行)*",
                )

        return "\n".join(parts)

    def _parse_docx(self, path: Path) -> str:
        """Extract paragraph text from a docx file."""
        from docx import Document

        doc = Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs) if paragraphs else "(no text found)"

    def _parse_pdf(self, path: Path) -> str:
        """Extract text from PDF pages."""
        import pdfplumber

        parts: list[str] = []
        with pdfplumber.open(str(path)) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if text:
                    parts.append(f"## Page {i}\n\n{text}")
        return "\n\n".join(parts) if parts else "(no extractable text)"

    async def _parse_json(self, path: Path) -> str:
        """Read and pretty-print JSON."""
        async with aiofiles.open(path, encoding="utf-8") as f:
            raw = await f.read()
        data = json.loads(raw)
        return json.dumps(data, ensure_ascii=False, indent=2)

    async def _parse_csv(self, path: Path) -> str:
        """Read CSV as Markdown table."""
        async with aiofiles.open(path, encoding="utf-8", newline="") as f:
            content = await f.read()
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        if not rows:
            return "(empty CSV)"

        # Pad columns for uniformity
        col_count = max(len(r) for r in rows)
        header = list(rows[0]) + [""] * (col_count - len(rows[0]))
        parts = ["| " + " | ".join(header) + " |"]
        parts.append("|" + "|".join(["---"] * col_count) + "|")
        for row in rows[1 : self._max_table_rows + 1]:
            padded = list(row) + [""] * (col_count - len(row))
            parts.append("| " + " | ".join(padded) + " |")
        return "\n".join(parts)

    async def _parse_yaml(self, path: Path) -> str:
        """Read YAML as structured text."""
        import yaml

        async with aiofiles.open(path, encoding="utf-8") as f:
            raw = await f.read()
        data = yaml.safe_load(raw)
        return yaml.dump(data, allow_unicode=True, default_flow_style=False)

    async def _parse_text(self, path: Path) -> str:
        """Default text reader (UTF-8)."""
        async with aiofiles.open(
            path,
            encoding="utf-8",
            errors="replace",
        ) as f:
            return await f.read()
